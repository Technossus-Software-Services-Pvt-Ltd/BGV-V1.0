"""Tests for app.services.batch.parser module."""

import pytest
from app.services.batch.parser import (
    parse_csv, parse_excel, parse_import_file,
    ParseError, ParsedCandidate,
    _resolve_column, _build_column_map, _parse_rows,
)


class TestResolveColumn:
    def test_exact_match(self):
        assert _resolve_column("candidate_id") == "candidate_id"
        assert _resolve_column("name") == "name"
        assert _resolve_column("email") == "email"

    def test_alias_match(self):
        assert _resolve_column("Full Name") == "name"
        assert _resolve_column("Candidate Name") == "name"
        assert _resolve_column("Email Address") == "email"
        assert _resolve_column("Phone Number") == "phone"
        assert _resolve_column("Date of Birth") == "dob"

    def test_case_insensitive(self):
        assert _resolve_column("CANDIDATE_ID") == "candidate_id"
        assert _resolve_column("NAME") == "name"

    def test_unknown_column(self):
        assert _resolve_column("unknown_col") is None
        assert _resolve_column("address") is None


class TestBuildColumnMap:
    def test_all_required_present(self):
        headers = ["candidate_id", "name", "email"]
        column_map, missing = _build_column_map(headers)
        assert "candidate_id" in column_map
        assert "name" in column_map
        assert missing == []

    def test_missing_required(self):
        headers = ["email", "phone"]
        column_map, missing = _build_column_map(headers)
        assert "candidate_id" in missing or "name" in missing

    def test_aliases_mapped(self):
        headers = ["Full Name", "CandidateId", "Mobile"]
        column_map, missing = _build_column_map(headers)
        assert column_map.get("name") == "Full Name"
        assert column_map.get("candidate_id") == "CandidateId"
        assert column_map.get("phone") == "Mobile"
        assert missing == []


class TestParseRows:
    def test_valid_rows(self):
        rows = [
            {"candidate_id": "C001", "name": "Priya Sharma", "email": "priya@test.com"},
            {"candidate_id": "C002", "name": "Rahul Kumar", "email": ""},
        ]
        column_map = {"candidate_id": "candidate_id", "name": "name", "email": "email"}
        candidates, errors = _parse_rows(rows, column_map)
        assert len(candidates) == 2
        assert candidates[0].candidate_id == "C001"
        assert candidates[0].name == "Priya Sharma"
        assert candidates[1].email is None  # empty string -> None

    def test_missing_candidate_id(self):
        rows = [{"candidate_id": "", "name": "Priya"}]
        column_map = {"candidate_id": "candidate_id", "name": "name"}
        candidates, errors = _parse_rows(rows, column_map)
        assert len(candidates) == 0
        assert len(errors) == 1
        assert "Missing candidate_id" in errors[0]

    def test_missing_name(self):
        rows = [{"candidate_id": "C001", "name": ""}]
        column_map = {"candidate_id": "candidate_id", "name": "name"}
        candidates, errors = _parse_rows(rows, column_map)
        assert len(candidates) == 0
        assert "Missing name" in errors[0]

    def test_too_long_candidate_id(self):
        rows = [{"candidate_id": "x" * 101, "name": "Test"}]
        column_map = {"candidate_id": "candidate_id", "name": "name"}
        candidates, errors = _parse_rows(rows, column_map)
        assert len(candidates) == 0
        assert "exceeds 100" in errors[0]

    def test_invalid_email(self):
        rows = [{"candidate_id": "C001", "name": "Test", "email": "not-an-email"}]
        column_map = {"candidate_id": "candidate_id", "name": "name", "email": "email"}
        candidates, errors = _parse_rows(rows, column_map)
        assert len(candidates) == 0
        assert "Invalid email" in errors[0]


class TestParseCSV:
    def test_valid_csv(self):
        csv_content = b"candidate_id,name,email\nC001,Priya Sharma,priya@test.com\nC002,Rahul Kumar,rahul@test.com\n"
        candidates, errors = parse_csv(csv_content)
        assert len(candidates) == 2
        assert candidates[0].candidate_id == "C001"
        assert candidates[0].name == "Priya Sharma"

    def test_csv_with_bom(self):
        csv_content = b"\xef\xbb\xbfcandidate_id,name\nC001,Test User\n"
        candidates, errors = parse_csv(csv_content)
        assert len(candidates) == 1

    def test_csv_no_headers_raises(self):
        with pytest.raises(ParseError, match="no headers"):
            parse_csv(b"")

    def test_csv_no_data_raises(self):
        with pytest.raises(ParseError, match="no data"):
            parse_csv(b"candidate_id,name\n")

    def test_csv_missing_required_columns_raises(self):
        with pytest.raises(ParseError, match="Missing required"):
            parse_csv(b"email,phone\ntest@x.com,123\n")

    def test_csv_too_many_rows_raises(self):
        header = b"candidate_id,name\n"
        rows = b"".join(f"C{i:04d},User {i}\n".encode() for i in range(5001))
        with pytest.raises(ParseError, match="Maximum allowed is 5000"):
            parse_csv(header + rows)

    def test_csv_latin1_fallback(self):
        # Latin-1 encoded content
        content = "candidate_id,name\nC001,Jos\xe9 García\n".encode("latin-1")
        candidates, errors = parse_csv(content)
        assert len(candidates) == 1

    def test_csv_with_aliases(self):
        csv_content = b"Full Name,CandidateId,Email Address\nPriya,C001,p@t.com\n"
        candidates, errors = parse_csv(csv_content)
        assert len(candidates) == 1
        assert candidates[0].name == "Priya"


class TestParseExcel:
    def test_valid_xlsx(self):
        """Test parsing a valid xlsx file."""
        try:
            from openpyxl import Workbook
            import io

            wb = Workbook()
            ws = wb.active
            ws.append(["candidate_id", "name", "email"])
            ws.append(["C001", "Priya Sharma", "priya@test.com"])
            ws.append(["C002", "Rahul Kumar", "rahul@test.com"])

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            candidates, errors = parse_excel(buffer.read())
            assert len(candidates) == 2
            assert candidates[0].candidate_id == "C001"
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_xlsx_missing_required_columns(self):
        try:
            from openpyxl import Workbook
            import io

            wb = Workbook()
            ws = wb.active
            ws.append(["email", "phone"])
            ws.append(["a@b.com", "123"])

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            with pytest.raises(ParseError, match="Missing required"):
                parse_excel(buffer.read())
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_xlsx_no_data_rows(self):
        try:
            from openpyxl import Workbook
            import io

            wb = Workbook()
            ws = wb.active
            ws.append(["candidate_id", "name"])

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            with pytest.raises(ParseError, match="no data"):
                parse_excel(buffer.read())
        except ImportError:
            pytest.skip("openpyxl not installed")


class TestParseImportFile:
    def test_csv_file(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_bytes(b"candidate_id,name\nC001,Test\n")
        candidates, errors = parse_import_file(str(csv_file), "test.csv")
        assert len(candidates) == 1

    def test_xlsx_file(self, tmp_path):
        try:
            from openpyxl import Workbook
            import io

            wb = Workbook()
            ws = wb.active
            ws.append(["candidate_id", "name"])
            ws.append(["C001", "Test User"])

            xlsx_file = tmp_path / "test.xlsx"
            wb.save(str(xlsx_file))

            candidates, errors = parse_import_file(str(xlsx_file), "test.xlsx")
            assert len(candidates) == 1
        except ImportError:
            pytest.skip("openpyxl not installed")
