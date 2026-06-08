import csv
import io
from pathlib import Path
from typing import List, Tuple

from app.core.logging import get_logger

logger = get_logger("batch.parser")

# Required columns (case-insensitive matching)
REQUIRED_COLUMNS = {"candidate_id", "name"}
OPTIONAL_COLUMNS = {"email", "phone", "dob", "gender"}
ALL_KNOWN_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS

# Column name aliases (map common variants to canonical names)
COLUMN_ALIASES = {
    "candidate_id": "candidate_id",
    "candidateid": "candidate_id",
    "cand_id": "candidate_id",
    "id": "candidate_id",
    "name": "name",
    "full_name": "name",
    "fullname": "name",
    "candidate_name": "name",
    "candidatename": "name",
    "email": "email",
    "email_address": "email",
    "emailaddress": "email",
    "phone": "phone",
    "phone_number": "phone",
    "phonenumber": "phone",
    "mobile": "phone",
    "dob": "dob",
    "date_of_birth": "dob",
    "dateofbirth": "dob",
    "birth_date": "dob",
    "birthdate": "dob",
    "gender": "gender",
    "sex": "gender",
}


from app.core.exceptions import BatchParseError


class ParseError(BatchParseError):
    """Raised when file parsing fails. Inherits from BatchParseError (→ HTTP 400)."""
    pass


class ParsedCandidate:
    """A single candidate row parsed from the import file."""

    __slots__ = ("row_number", "candidate_id", "name", "email", "phone", "dob", "gender")

    def __init__(
        self,
        row_number: int,
        candidate_id: str,
        name: str,
        email: str | None = None,
        phone: str | None = None,
        dob: str | None = None,
        gender: str | None = None,
    ):
        self.row_number = row_number
        self.candidate_id = candidate_id
        self.name = name
        self.email = email
        self.phone = phone
        self.dob = dob
        self.gender = gender


def _resolve_column(raw_header: str) -> str | None:
    """Map a raw header string to a canonical column name, or None if unrecognised."""
    key = raw_header.strip().lower().replace(" ", "_").replace("-", "_")
    return COLUMN_ALIASES.get(key)


def _parse_rows(rows: List[dict], column_map: dict[str, str]) -> Tuple[List[ParsedCandidate], List[str]]:
    """Validate and convert raw dicts into ParsedCandidate objects."""
    candidates: List[ParsedCandidate] = []
    errors: List[str] = []

    for idx, row in enumerate(rows, start=2):  # Row 1 is header
        raw_cid = row.get(column_map.get("candidate_id", ""), "").strip()
        raw_name = row.get(column_map.get("name", ""), "").strip()

        if not raw_cid:
            errors.append(f"Row {idx}: Missing candidate_id")
            continue
        if not raw_name:
            errors.append(f"Row {idx}: Missing name")
            continue

        # Length validation
        if len(raw_cid) > 100:
            errors.append(f"Row {idx}: candidate_id exceeds 100 characters")
            continue
        if len(raw_name) > 255:
            errors.append(f"Row {idx}: name exceeds 255 characters")
            continue

        email = row.get(column_map.get("email", ""), "").strip() or None
        phone = row.get(column_map.get("phone", ""), "").strip() or None
        dob = row.get(column_map.get("dob", ""), "").strip() or None
        gender = row.get(column_map.get("gender", ""), "").strip() or None

        # Basic email format check
        if email and "@" not in email:
            errors.append(f"Row {idx}: Invalid email format '{email}'")
            continue

        candidates.append(
            ParsedCandidate(
                row_number=idx,
                candidate_id=raw_cid,
                name=raw_name,
                email=email,
                phone=phone,
                dob=dob,
                gender=gender,
            )
        )

    return candidates, errors


def _build_column_map(headers: List[str]) -> Tuple[dict[str, str], List[str]]:
    """Build mapping from canonical column name to actual header string.
    Returns (column_map, missing_required).
    """
    column_map: dict[str, str] = {}  # canonical -> raw header
    for raw in headers:
        canonical = _resolve_column(raw)
        if canonical and canonical not in column_map:
            column_map[canonical] = raw

    missing = [col for col in REQUIRED_COLUMNS if col not in column_map]
    return column_map, missing


def parse_csv(file_bytes: bytes) -> Tuple[List[ParsedCandidate], List[str]]:
    """Parse a CSV file and return (candidates, errors)."""
    try:
        text = file_bytes.decode("utf-8-sig")  # Handle BOM
    except UnicodeDecodeError:
        try:
            text = file_bytes.decode("latin-1")
        except UnicodeDecodeError:
            raise ParseError("Unable to decode file. Please use UTF-8 encoding.")

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ParseError("CSV file has no headers")

    column_map, missing = _build_column_map(list(reader.fieldnames))
    if missing:
        raise ParseError(f"Missing required columns: {', '.join(missing)}. Found: {', '.join(reader.fieldnames)}")

    rows = list(reader)
    if not rows:
        raise ParseError("CSV file has no data rows")

    if len(rows) > 5000:
        raise ParseError(f"CSV contains {len(rows)} rows. Maximum allowed is 5000.")

    logger.info("csv_parsed", total_rows=len(rows), columns=list(column_map.keys()))
    return _parse_rows(rows, column_map)


def parse_excel(file_bytes: bytes) -> Tuple[List[ParsedCandidate], List[str]]:
    """Parse an Excel (.xlsx) file and return (candidates, errors)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ParseError("openpyxl is required for Excel parsing. Install it with: pip install openpyxl")

    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ParseError(f"Invalid Excel file: {e}")

    ws = wb.active
    if ws is None:
        raise ParseError("Excel file has no active sheet")

    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ParseError("Excel file is empty")

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    headers = [h for h in headers if h]  # Remove empty trailing columns
    if not headers:
        raise ParseError("Excel file has no headers in the first row")

    column_map, missing = _build_column_map(headers)
    if missing:
        raise ParseError(f"Missing required columns: {', '.join(missing)}. Found: {', '.join(headers)}")

    # Read all data rows into dicts
    data_rows: List[dict] = []
    for row_values in rows_iter:
        if all(cell is None or str(cell).strip() == "" for cell in row_values):
            continue  # Skip entirely blank rows
        row_dict = {}
        for i, header in enumerate(headers):
            val = row_values[i] if i < len(row_values) else None
            row_dict[header] = str(val).strip() if val is not None else ""
        data_rows.append(row_dict)

    wb.close()

    if not data_rows:
        raise ParseError("Excel file has no data rows")

    if len(data_rows) > 5000:
        raise ParseError(f"Excel file contains {len(data_rows)} rows. Maximum allowed is 5000.")

    logger.info("excel_parsed", total_rows=len(data_rows), columns=list(column_map.keys()))
    return _parse_rows(data_rows, column_map)


def parse_import_file(file_path: str, original_filename: str) -> Tuple[List[ParsedCandidate], List[str]]:
    """Detect file type and parse accordingly."""
    file_bytes = Path(file_path).read_bytes()
    ext = Path(original_filename).suffix.lower()

    if ext == ".csv":
        return parse_csv(file_bytes)
    elif ext in (".xlsx", ".xls"):
        return parse_excel(file_bytes)
    else:
        raise ParseError(f"Unsupported file format: {ext}. Use .csv or .xlsx")
